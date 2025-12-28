import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict


ARMOR_TYPE_COLORS = {
    'Helmet': 'FFC7CE',
    'Gauntlets': 'C6EFCE',
    'Chest Armor': 'FFEB9C',
    'Leg Armor': 'D9E1F2',
    'Hunter Cloak': 'F4CCCC',
    'Warlock Bond': 'CCE5FF',
    'Titan Mark': 'EAD1DC'
}

DUPLICATE_HIGHLIGHT_COLOR = 'FFD7B5'


def generate_summary_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Generate summary statistics per class.

    Args:
        df: Processed armor DataFrame

    Returns:
        Summary DataFrame with metrics per class
    """
    summary_rows = []

    for class_name in ['Hunter', 'Warlock', 'Titan']:
        class_df = df[df['Class'] == class_name]

        if len(class_df) == 0:
            continue

        total_pieces = len(class_df)
        avg_total = class_df['Total Stat'].mean()
        highest_total = class_df['Total Stat'].max()
        lowest_total = class_df['Total Stat'].min()

        archetype_counts = class_df['Archetype'].value_counts().to_dict()
        archetype_breakdown = ', '.join([f"{k}: {v}" for k, v in sorted(archetype_counts.items())])

        duplicate_count = class_df['Duplicate Same Archetype+Tertiary Across Sets'].sum()

        summary_rows.append({
            'Class': class_name,
            'Total Pieces': total_pieces,
            'Average Total Stat': round(avg_total, 2),
            'Highest Total Stat': highest_total,
            'Lowest Total Stat': lowest_total,
            'Archetype Breakdown': archetype_breakdown,
            'Duplicate Count': int(duplicate_count)
        })

    return pd.DataFrame(summary_rows)


def apply_sheet_styling(ws, df: pd.DataFrame, class_name: str = None):
    """
    Apply styling to a worksheet including colors, bold columns, and filters.

    Args:
        ws: Openpyxl worksheet object
        df: DataFrame containing the data
        class_name: Optional class name for class-specific sheets
    """
    bold_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = bold_font

    ws.auto_filter.ref = ws.dimensions

    if class_name:
        header_row = list(df.columns)
        archetype_col_idx = header_row.index('Archetype') + 1 if 'Archetype' in header_row else None
        armor_set_col_idx = header_row.index('Armor Set') + 1 if 'Armor Set' in header_row else None
        duplicate_col_idx = header_row.index('Duplicate Same Archetype+Tertiary Across Sets') + 1 if 'Duplicate Same Archetype+Tertiary Across Sets' in header_row else None
        exact_duplicate_col_idx = header_row.index('Exact Duplicate (Including Tuning)') + 1 if 'Exact Duplicate (Including Tuning)' in header_row else None
        armor_type_col_idx = header_row.index('Armor Type') + 1 if 'Armor Type' in header_row else None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df) + 1), start=0):
            if row_idx >= len(df):
                break

            armor_type = df.iloc[row_idx]['Armor Type']

            if armor_type in ARMOR_TYPE_COLORS:
                fill_color = ARMOR_TYPE_COLORS[armor_type]
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

                for cell in row:
                    cell.fill = fill

            if archetype_col_idx:
                row[archetype_col_idx - 1].font = bold_font
            if armor_set_col_idx:
                row[armor_set_col_idx - 1].font = bold_font

            if duplicate_col_idx and df.iloc[row_idx]['Duplicate Same Archetype+Tertiary Across Sets']:
                dup_fill = PatternFill(start_color=DUPLICATE_HIGHLIGHT_COLOR,
                                      end_color=DUPLICATE_HIGHLIGHT_COLOR,
                                      fill_type='solid')
                row[duplicate_col_idx - 1].fill = dup_fill

            if exact_duplicate_col_idx and df.iloc[row_idx]['Exact Duplicate (Including Tuning)']:
                exact_dup_fill = PatternFill(start_color='ED985F',  # Brighter orange for exact duplicates
                                            end_color='ED985F',
                                            fill_type='solid')
                row[exact_duplicate_col_idx - 1].fill = exact_dup_fill
                row[exact_duplicate_col_idx - 1].font = bold_font

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ''
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(df: pd.DataFrame, output_path: str) -> str:
    """
    Export processed armor data to styled Excel workbook.

    Args:
        df: Processed DataFrame with armor data
        output_path: Path where Excel file should be saved

    Returns:
        Path to the created Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)

    summary_df = generate_summary_data(df)
    ws_summary = wb.create_sheet('Summary')
    for row in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(row)
    apply_sheet_styling(ws_summary, summary_df)

    for class_name in ['Hunter', 'Warlock', 'Titan']:
        class_df = df[df['Class'] == class_name].reset_index(drop=True)

        if len(class_df) > 0:
            ws = wb.create_sheet(class_name)

            for row in dataframe_to_rows(class_df, index=False, header=True):
                ws.append(row)

            apply_sheet_styling(ws, class_df, class_name)

    wb.save(output_path)
    return output_path

